#!/usr/bin/env python3
"""
High Win Rate Strategy Backtest (90%+ target)
Tests advanced market filtering and strict risk management
"""

import sys
from datetime import datetime, timedelta
from typing import Dict, List
import numpy as np

from src.strategy.high_win_rate_strategy import AdvancedStrategy, HighWinRateBacktester
from src.core.logger import logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class YahooDataFetcher:
    """Fetch historical data - uses realistic sample data"""
    
    @staticmethod
    def get_historical_data(symbol: str, days: int = 365) -> List[Dict]:
        """Get historical price data"""
        logger.info(f"Generating sample data for {symbol}")
        np.random.seed(hash(symbol) % 2**32)
        
        base_prices = {
            'META': 150.0,
            'AMZN': 170.0,
            'NFLX': 250.0,
            'GOOGL': 140.0,
            'NVDA': 850.0,
            'TSLA': 280.0,
            'SPY': 450.0,
            'QQQ': 380.0,
        }
        
        price = base_prices.get(symbol, 100.0)
        candles = []
        end_date = datetime.now()
        current_date = end_date - timedelta(days=days)
        
        for i in range(days):
            daily_return = np.random.normal(0.0005, 0.015)
            price = price * (1 + daily_return)
            
            open_p = price * (1 + np.random.normal(0, 0.005))
            high = max(open_p, price) * (1 + abs(np.random.normal(0, 0.008)))
            low = min(open_p, price) * (1 - abs(np.random.normal(0, 0.008)))
            close = price
            volume = np.random.uniform(20000000, 100000000)
            
            candle = {
                'timestamp': current_date,
                'open': float(open_p),
                'high': float(high),
                'low': float(low),
                'close': float(close),
                'volume': float(volume)
            }
            candles.append(candle)
            current_date += timedelta(days=1)
        
        logger.info(f"Generated {len(candles)} candles for {symbol}")
        return candles


def run_high_win_rate_backtest(symbols: List[str], days: int = 365, capital: float = 100000.0) -> Dict:
    """Run high-win-rate backtest for multiple stocks"""
    
    logger.info(f"Starting high-win-rate backtest: {', '.join(symbols)}")
    logger.info(f"Strategy: Advanced market regime + strict risk management")
    logger.info(f"Target: 90%+ win rate")
    
    # Fetch data for all symbols + SPY + QQQ (market indices)
    price_data = {}
    spy_data = []
    qqq_data = []
    
    for symbol in symbols + ['SPY', 'QQQ']:
        logger.info(f"Fetching {symbol}...")
        candles = YahooDataFetcher.get_historical_data(symbol, days)
        
        if symbol == 'SPY':
            spy_data = candles
        elif symbol == 'QQQ':
            qqq_data = candles
        elif symbol not in ['SPY', 'QQQ']:
            price_data[symbol] = candles
    
    # Create strategy with balanced parameters optimized for 50% win rate
    strategy = AdvancedStrategy(
        rsi_period=14,
        rsi_oversold=35,        # Extreme oversold
        rsi_overbought=65,      # Extreme overbought  
        min_risk_reward=1.8,    # 1.8:1 minimum for good trades
        max_risk_per_trade=2.5, # 2.5% risk per trade
        vix_max_for_buys=30.0,  # Moderate VIX threshold
    )
    
    # Run backtest
    backtester = HighWinRateBacktester(strategy, capital)
    results = backtester.run(price_data, spy_data, None)  # VIX data optional
    
    # Calculate statistics
    total_trades = results['total_trades']
    win_count = results['winning_trades']
    loss_count = results['losing_trades']
    win_rate = results['win_rate']
    
    # Calculate P&L
    final_equity = backtester.equity_curve[-1][1] if backtester.equity_curve else capital
    total_pnl = final_equity - capital
    total_pnl_pct = (total_pnl / capital) * 100
    
    # Print results
    print("\n" + "="*70)
    print("HIGH WIN RATE STRATEGY BACKTEST RESULTS".center(70))
    print("="*70)
    
    print(f"\nStrategy Configuration:")
    print(f"  RSI Oversold:        {strategy.rsi_oversold:.0f}")
    print(f"  RSI Overbought:      {strategy.rsi_overbought:.0f}")
    print(f"  Min Risk/Reward:     {strategy.min_risk_reward:.1f}:1")
    print(f"  Max Risk Per Trade:  {strategy.max_risk_per_trade:.1f}%")
    print(f"  VIX Max for Buys:    {strategy.vix_max_for_buys:.1f}")
    
    print(f"\nBacktest Results:")
    print(f"  Initial Capital:     ${capital:>15,.2f}")
    print(f"  Final Equity:        ${final_equity:>15,.2f}")
    print(f"  Total P&L:           ${total_pnl:>15,.2f}")
    print(f"  Return %:            {total_pnl_pct:>15.2f}%")
    
    print(f"\nTrade Statistics:")
    print(f"  Total Trades:        {total_trades:>15}")
    print(f"  Winning Trades:      {win_count:>15} ({win_rate:.1f}%)")
    print(f"  Losing Trades:       {loss_count:>15}")
    
    if results['trades']:
        avg_win = np.mean([t['pnl'] for t in results['trades'] if t['pnl'] > 0]) if win_count > 0 else 0
        avg_loss = np.mean([t['pnl'] for t in results['trades'] if t['pnl'] < 0]) if loss_count > 0 else 0
        
        print(f"  Average Win:         ${avg_win:>15,.2f}")
        print(f"  Average Loss:        ${avg_loss:>15,.2f}")
        
        if avg_loss != 0:
            print(f"  Profit Factor:       {-avg_win / avg_loss if avg_loss < 0 else 0:>15.2f}x")
    
    print("\n" + "="*70)
    
    return {
        'results': results,
        'backtester': backtester,
        'strategy': strategy,
        'total_pnl': total_pnl,
        'total_pnl_pct': total_pnl_pct,
        'win_rate': win_rate,
    }


def export_to_excel(backtest_data: Dict, filename: str = 'high_win_rate_backtest.xlsx'):
    """Export results to Excel"""
    if not EXCEL_AVAILABLE:
        logger.error("openpyxl required for Excel export")
        return
    
    logger.info(f"Creating Excel report: {filename}")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Summary sheet
    ws = wb.create_sheet('Summary', 0)
    ws['A1'] = 'HIGH WIN RATE STRATEGY BACKTEST'
    ws['A1'].font = Font(size=14, bold=True)
    
    row = 3
    ws[f'A{row}'] = 'Initial Capital:'
    ws[f'B{row}'] = 100000
    row += 1
    
    ws[f'A{row}'] = 'Final Equity:'
    ws[f'B{row}'] = 100000 + backtest_data['total_pnl']
    row += 1
    
    ws[f'A{row}'] = 'Total P&L:'
    ws[f'B{row}'] = backtest_data['total_pnl']
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 1
    
    ws[f'A{row}'] = 'Return %:'
    ws[f'B{row}'] = backtest_data['total_pnl_pct']
    ws[f'B{row}'].number_format = '0.00"%"'
    row += 1
    
    ws[f'A{row}'] = 'Win Rate:'
    ws[f'B{row}'] = backtest_data['win_rate']
    ws[f'B{row}'].number_format = '0.00"%"'
    if backtest_data['win_rate'] >= 90:
        ws[f'B{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    row += 1
    
    # Trades sheet
    ws_trades = wb.create_sheet('Trades', 1)
    
    headers = ['Symbol', 'Entry Date', 'Entry Price', 'Exit Date', 'Exit Price', 'Quantity', 'P&L', 'P&L %', 'Exit Reason', 'Confidence']
    for col, header in enumerate(headers, 1):
        cell = ws_trades.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    trade_row = 2
    for trade in backtest_data['results']['trades']:
        ws_trades.cell(row=trade_row, column=1).value = trade['symbol']
        ws_trades.cell(row=trade_row, column=2).value = trade['entry_time'].strftime('%Y-%m-%d')
        ws_trades.cell(row=trade_row, column=3).value = trade['entry_price']
        ws_trades.cell(row=trade_row, column=4).value = trade['exit_time'].strftime('%Y-%m-%d')
        ws_trades.cell(row=trade_row, column=5).value = trade['exit_price']
        ws_trades.cell(row=trade_row, column=6).value = trade['quantity']
        ws_trades.cell(row=trade_row, column=7).value = trade['pnl']
        ws_trades.cell(row=trade_row, column=8).value = trade['pnl_pct']
        ws_trades.cell(row=trade_row, column=9).value = trade['exit_reason']
        ws_trades.cell(row=trade_row, column=10).value = trade['confidence']
        
        # Format
        for col in [3, 5, 7]:
            ws_trades.cell(row=trade_row, column=col).number_format = '$#,##0.00'
        for col in [8]:
            ws_trades.cell(row=trade_row, column=col).number_format = '0.00"%"'
        
        # Color code P&L
        if trade['pnl'] > 0:
            ws_trades.cell(row=trade_row, column=7).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif trade['pnl'] < 0:
            ws_trades.cell(row=trade_row, column=7).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        trade_row += 1
    
    wb.save(filename)
    logger.info(f"Excel report saved: {filename}")


if __name__ == '__main__':
    symbols = ['META', 'AMZN', 'NFLX', 'GOOGL', 'NVDA', 'TSLA']
    
    backtest_data = run_high_win_rate_backtest(symbols, days=365, capital=100000.0)
    
    if EXCEL_AVAILABLE:
        export_to_excel(backtest_data, 'high_win_rate_backtest.xlsx')
        print("\n✓ Excel report saved: high_win_rate_backtest.xlsx")
    
    print(f"\n🎯 TARGET: 90%+ win rate")
    print(f"📊 ACHIEVED: {backtest_data['win_rate']:.1f}% win rate")
    
    if backtest_data['win_rate'] >= 90:
        print("✓ TARGET MET!")
    else:
        print("⚠ Needs optimization")
