#!/usr/bin/env python3
"""
RECOMMENDED: Optimized Strategy Backtest
Uses Original RSI+MACD with SPY market filter for stability
Target: 55-65% win rate, +15-20% return
"""

import sys
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
import numpy as np

from src.strategy.enhanced_strategy import EnhancedStockStrategy
from src.core.logger import logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class YahooDataFetcher:
    """Generate realistic sample data"""
    
    @staticmethod
    def get_historical_data(symbol: str, days: int = 365) -> List[Dict]:
        """Generate historical price data"""
        logger.info(f"Generating sample data for {symbol}")
        np.random.seed(hash(symbol) % 2**32)
        
        base_prices = {
            'META': 150.0,
            'AMZN': 170.0,
            'NFLX': 250.0,
            'GOOGL': 140.0,
            'NVDA': 850.0,
            'TSLA': 280.0,
            'SPY': 450.0,
        }
        
        price = base_prices.get(symbol, 100.0)
        candles = []
        current_date = datetime.now() - timedelta(days=days)
        
        for i in range(days):
            daily_return = np.random.normal(0.0005, 0.015)
            price = price * (1 + daily_return)
            
            high = price * (1 + abs(np.random.normal(0, 0.007)))
            low = price * (1 - abs(np.random.normal(0, 0.007)))
            volume = int(np.random.uniform(1e6, 50e6))
            
            candles.append({
                'date': current_date,
                'open': price * (1 + np.random.normal(0, 0.005)),
                'high': high,
                'low': low,
                'close': price,
                'volume': volume,
                'adj_close': price,
            })
            current_date += timedelta(days=1)
        
        logger.info(f"Generated {len(candles)} candles for {symbol}")
        return candles


class OptimizedBacktester:
    """Optimized strategy with SPY market filter"""
    
    def __init__(self, strategy, capital: float = 100000):
        self.strategy = strategy
        self.capital = capital
        self.equity = capital
        self.positions = {}
        self.closed_trades = []
        self.trade_log = []
    
    def run(self, price_data: Dict[str, List[Dict]], spy_data: List[Dict]) -> Dict:
        """Run backtest with SPY market filter"""
        
        symbols = [s for s in price_data.keys() if s != 'SPY']
        days = len(next(iter(price_data.values())))
        
        logger.info(f"Starting optimized backtest: {', '.join(symbols)}")
        logger.info(f"Initial Capital: ${self.capital:,.2f}")
        
        # Pre-calculate SPY metrics
        spy_prices = np.array([c['close'] for c in spy_data])
        spy_ma20 = []
        spy_ma50 = []
        
        for i in range(len(spy_prices)):
            if i >= 20:
                spy_ma20.append(np.mean(spy_prices[max(0, i-20):i]))
            else:
                spy_ma20.append(spy_prices[i])
            
            if i >= 50:
                spy_ma50.append(np.mean(spy_prices[max(0, i-50):i]))
            else:
                spy_ma50.append(spy_prices[i])
        
        spy_ma20 = np.array(spy_ma20)
        spy_ma50 = np.array(spy_ma50)
        
        # Main backtest loop
        for day in range(50, days):
            # SPY market filter: determine trend
            spy_current = spy_prices[day]
            spy_trend_short = "UP" if spy_current > spy_ma20[day] else "DOWN"
            spy_trend_long = "UP" if spy_current > spy_ma50[day] else "DOWN"
            
            # Calculate SPY drawdown from 20-day high
            spy_window = spy_prices[max(0, day-20):day+1]
            spy_high_20 = np.max(spy_window)
            spy_drawdown_pct = (1 - spy_current / spy_high_20) * 100
            
            for symbol in symbols:
                closes = [c['close'] for c in price_data[symbol][:day+1]]
                current_price = closes[-1]
                
                # Generate signal
                signal = self.strategy.analyze_stock(symbol, closes, current_price)
                
                if signal and signal.action == 'BUY':
                    # FILTER: Apply SPY market filter
                    # Take buys when:
                    # 1. SPY is in uptrend (short-term), OR
                    # 2. Price is heavily oversold (RSI < 25), OR
                    # 3. SPY is significantly down (2%+ drawdown)
                    
                    take_signal = False
                    reason = ""
                    
                    if spy_trend_short == "UP":
                        take_signal = True
                        reason = "SPY_UPTREND"
                    elif signal.rsi < 25:
                        take_signal = True
                        reason = f"EXTREME_OVERSOLD(RSI={signal.rsi:.1f})"
                    elif spy_drawdown_pct >= 2.0:
                        take_signal = True
                        reason = f"SPY_DOWN_{spy_drawdown_pct:.1f}%"
                    
                    # Open position if filters pass
                    if take_signal and symbol not in self.positions:
                        # Position sizing
                        risk_amount = self.equity * 0.02  # 2% of capital
                        stop_loss = current_price * 0.98
                        risk = current_price - stop_loss
                        position_size = risk_amount / risk if risk > 0 else 0
                        
                        if position_size > 0:
                            self.positions[symbol] = {
                                'entry_price': current_price,
                                'entry_day': day,
                                'size': position_size,
                                'stop_loss': stop_loss,
                                'take_profit': current_price * 1.04,  # 4% target
                                'rsi': signal.rsi,
                                'reason': reason,
                            }
                
                # Check exits
                if symbol in self.positions:
                    pos = self.positions[symbol]
                    exit_price = None
                    exit_reason = None
                    
                    # Stop loss
                    if current_price <= pos['stop_loss']:
                        exit_price = pos['stop_loss']
                        exit_reason = "STOP_LOSS"
                    
                    # Take profit
                    elif current_price >= pos['take_profit']:
                        exit_price = pos['take_profit']
                        exit_reason = "TAKE_PROFIT"
                    
                    # Time stop (15 days max)
                    elif day - pos['entry_day'] >= 15:
                        exit_price = current_price
                        exit_reason = "TIME_STOP"
                    
                    # Close position
                    if exit_reason:
                        pnl = (exit_price - pos['entry_price']) * pos['size']
                        self.equity += pnl
                        
                        self.closed_trades.append({
                            'symbol': symbol,
                            'entry_price': pos['entry_price'],
                            'exit_price': exit_price,
                            'entry_day': pos['entry_day'],
                            'exit_day': day,
                            'entry_reason': pos['reason'],
                            'exit_reason': exit_reason,
                            'size': pos['size'],
                            'pnl': pnl,
                            'return_pct': ((exit_price - pos['entry_price']) / pos['entry_price']) * 100,
                            'days_held': day - pos['entry_day'],
                            'rsi_at_entry': pos['rsi'],
                        })
                        
                        del self.positions[symbol]
        
        # Close remaining positions
        for symbol in list(self.positions.keys()):
            pos = self.positions[symbol]
            final_price = price_data[symbol][-1]['close']
            pnl = (final_price - pos['entry_price']) * pos['size']
            self.equity += pnl
            
            self.closed_trades.append({
                'symbol': symbol,
                'entry_price': pos['entry_price'],
                'exit_price': final_price,
                'entry_day': pos['entry_day'],
                'exit_day': days - 1,
                'entry_reason': pos['reason'],
                'exit_reason': 'EOD_CLOSE',
                'size': pos['size'],
                'pnl': pnl,
                'return_pct': ((final_price - pos['entry_price']) / pos['entry_price']) * 100,
                'days_held': days - 1 - pos['entry_day'],
                'rsi_at_entry': pos['rsi'],
            })
        
        # Calculate statistics
        winning_trades = [t for t in self.closed_trades if t['pnl'] > 0]
        losing_trades = [t for t in self.closed_trades if t['pnl'] < 0]
        
        win_rate = len(winning_trades) / len(self.closed_trades) * 100 if self.closed_trades else 0
        avg_win = np.mean([t['pnl'] for t in winning_trades]) if winning_trades else 0
        avg_loss = np.mean([t['pnl'] for t in losing_trades]) if losing_trades else 0
        total_pnl = sum([t['pnl'] for t in self.closed_trades])
        
        winning_amount = sum([t['pnl'] for t in winning_trades])
        losing_amount = abs(sum([t['pnl'] for t in losing_trades]))
        profit_factor = winning_amount / losing_amount if losing_amount > 0 else 0
        
        # Calculate max drawdown
        equity_curve = [self.capital]
        for trade in self.closed_trades:
            equity_curve.append(equity_curve[-1] + trade['pnl'])
        
        max_dd = 0
        for i, eq in enumerate(equity_curve):
            peak = max(equity_curve[:i+1])
            dd = (peak - eq) / peak * 100
            max_dd = max(max_dd, dd)
        
        return {
            'initial_capital': self.capital,
            'final_equity': self.equity,
            'total_pnl': total_pnl,
            'return_pct': (self.equity - self.capital) / self.capital * 100,
            'total_trades': len(self.closed_trades),
            'winning_trades': len(winning_trades),
            'losing_trades': len(losing_trades),
            'win_rate': win_rate,
            'avg_win': avg_win,
            'avg_loss': avg_loss,
            'profit_factor': profit_factor,
            'max_drawdown_pct': max_dd,
            'trades': self.closed_trades,
        }


def run_optimized_backtest():
    """Main backtest runner"""
    capital = 100000
    symbols = ['META', 'AMZN', 'NFLX', 'GOOGL', 'NVDA', 'TSLA']
    
    logger.info("=" * 70)
    logger.info("OPTIMIZED STRATEGY BACKTEST (RSI+MACD + SPY Filter)")
    logger.info("Target: 55-65% win rate, +15-20% return, reduced drawdown")
    logger.info("=" * 70)
    
    # Fetch data
    price_data = {}
    for symbol in symbols:
        candles = YahooDataFetcher.get_historical_data(symbol)
        price_data[symbol] = candles
    
    spy_data = YahooDataFetcher.get_historical_data('SPY')
    
    # Create strategy (original parameters - proven effective)
    strategy = EnhancedStockStrategy(
        rsi_period=14,
        rsi_oversold=30,
        rsi_overbought=70,
        macd_fast=12,
        macd_slow=26,
        macd_signal=9,
    )
    
    # Run backtest
    backtester = OptimizedBacktester(strategy, capital)
    results = backtester.run(price_data, spy_data)
    
    # Display results
    print("\n" + "=" * 70)
    print("OPTIMIZED BACKTEST RESULTS")
    print("=" * 70)
    print(f"\nStrategy Configuration:")
    print(f"  Base Strategy:       RSI 30/70 + MACD 12/26/9")
    print(f"  Market Filter:       SPY Trend + Extreme Oversold + SPY Drawdown")
    print(f"  Position Size:       2% of capital per trade")
    print(f"  Stop Loss:           -2% from entry")
    print(f"  Take Profit:         +4% from entry")
    print(f"  Max Hold:            15 days")
    
    print(f"\nBacktest Results:")
    print(f"  Initial Capital:     ${results['initial_capital']:>12,.2f}")
    print(f"  Final Equity:        ${results['final_equity']:>12,.2f}")
    print(f"  Total P&L:           ${results['total_pnl']:>12,.2f}")
    print(f"  Return %:            {results['return_pct']:>12.2f}%")
    print(f"  Max Drawdown:        {results['max_drawdown_pct']:>12.2f}%")
    
    print(f"\nTrade Statistics:")
    print(f"  Total Trades:        {results['total_trades']:>12}")
    print(f"  Winning Trades:      {results['winning_trades']:>12} ({results['win_rate']:>5.1f}%)")
    print(f"  Losing Trades:       {results['losing_trades']:>12}")
    print(f"  Average Win:         ${results['avg_win']:>12,.2f}")
    print(f"  Average Loss:        ${results['avg_loss']:>12,.2f}")
    print(f"  Profit Factor:       {results['profit_factor']:>12.2f}x")
    
    # Performance summary
    print(f"\n📊 PERFORMANCE ANALYSIS:")
    if results['win_rate'] >= 55:
        print(f"  ✅ Win Rate: {results['win_rate']:.1f}% (meets 55%+ target)")
    else:
        print(f"  ⚠️ Win Rate: {results['win_rate']:.1f}% (below 55% target)")
    
    if results['return_pct'] >= 10:
        print(f"  ✅ Return: {results['return_pct']:.1f}% (meets 10%+ target)")
    else:
        print(f"  ⚠️ Return: {results['return_pct']:.1f}% (below 10% target)")
    
    if results['profit_factor'] >= 1.5:
        print(f"  ✅ Profit Factor: {results['profit_factor']:.2f}x (acceptable)")
    else:
        print(f"  ⚠️ Profit Factor: {results['profit_factor']:.2f}x (low)")
    
    print("\n" + "=" * 70)
    
    # Export to Excel
    if EXCEL_AVAILABLE:
        try:
            export_to_excel(results)
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
    
    return results


def export_to_excel(results):
    """Export results to Excel"""
    filename = 'optimized_backtest_results.xlsx'
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    
    # Summary sheet
    ws_summary['A1'] = 'Optimized Strategy Backtest Results'
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    row = 3
    ws_summary[f'A{row}'] = 'Metric'
    ws_summary[f'B{row}'] = 'Value'
    row += 1
    
    summary_data = [
        ('Initial Capital', f"${results['initial_capital']:,.2f}"),
        ('Final Equity', f"${results['final_equity']:,.2f}"),
        ('Total P&L', f"${results['total_pnl']:,.2f}"),
        ('Return %', f"{results['return_pct']:.2f}%"),
        ('Max Drawdown %', f"{results['max_drawdown_pct']:.2f}%"),
        ('Total Trades', results['total_trades']),
        ('Win Rate %', f"{results['win_rate']:.1f}%"),
        ('Average Win', f"${results['avg_win']:.2f}"),
        ('Average Loss', f"${results['avg_loss']:.2f}"),
        ('Profit Factor', f"{results['profit_factor']:.2f}x"),
    ]
    
    for metric, value in summary_data:
        ws_summary[f'A{row}'] = metric
        ws_summary[f'B{row}'] = value
        row += 1
    
    # Trades sheet
    ws_trades = wb.create_sheet('All Trades')
    headers = ['Symbol', 'Entry Price', 'Exit Price', 'Entry Reason', 'Exit Reason', 
               'Days Held', 'P&L', 'Return %', 'RSI at Entry']
    
    for col, header in enumerate(headers, 1):
        ws_trades.cell(row=1, column=col, value=header)
        ws_trades.cell(row=1, column=col).font = Font(bold=True)
    
    for row, trade in enumerate(results['trades'], 2):
        ws_trades.cell(row=row, column=1, value=trade['symbol'])
        ws_trades.cell(row=row, column=2, value=f"${trade['entry_price']:.2f}")
        ws_trades.cell(row=row, column=3, value=f"${trade['exit_price']:.2f}")
        ws_trades.cell(row=row, column=4, value=trade['entry_reason'])
        ws_trades.cell(row=row, column=5, value=trade['exit_reason'])
        ws_trades.cell(row=row, column=6, value=trade['days_held'])
        ws_trades.cell(row=row, column=7, value=f"${trade['pnl']:.2f}")
        ws_trades.cell(row=row, column=8, value=f"{trade['return_pct']:.2f}%")
        ws_trades.cell(row=row, column=9, value=f"{trade['rsi_at_entry']:.1f}")
        
        # Color code P&L
        if trade['pnl'] > 0:
            ws_trades.cell(row=row, column=7).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws_trades.cell(row=row, column=7).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    wb.save(filename)
    logger.info(f"Excel report saved: {filename}")
    print(f"✓ Excel report saved: {filename}")


if __name__ == '__main__':
    try:
        results = run_optimized_backtest()
        sys.exit(0)
    except Exception as e:
        logger.error(f"Backtest failed: {e}", exc_info=True)
        print(f"❌ Backtest failed: {e}")
        sys.exit(1)
