#!/usr/bin/env python3
"""
Options Hedging Backtest
Compares trading strategy performance with protective puts
Shows how hedges reduce maximum losses while accepting hedge costs
"""

import sys
from datetime import datetime, timedelta
from typing import Dict, List
import numpy as np

from src.strategy.enhanced_strategy import EnhancedStockStrategy
from src.strategy.options_hedge import (
    OptionsHedgeManager,
    HedgeParameters,
    HedgeStrategy,
    HedgePosition
)
from src.core.logger import logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class YahooDataFetcher:
    """Generate realistic sample data"""
    
    @staticmethod
    def get_historical_data(symbol: str, days: int = 365) -> List[Dict]:
        """Generate historical price data"""
        np.random.seed(hash(symbol) % 2**32)
        
        base_prices = {
            'META': 150.0,
            'AMZN': 170.0,
            'NFLX': 250.0,
            'GOOGL': 140.0,
            'NVDA': 850.0,
            'TSLA': 280.0,
            'SPY': 450.0,
        }
        
        price = base_prices.get(symbol, 100.0)
        candles = []
        current_date = datetime.now() - timedelta(days=days)
        
        for i in range(days):
            daily_return = np.random.normal(0.0005, 0.015)
            price = price * (1 + daily_return)
            
            high = price * (1 + abs(np.random.normal(0, 0.007)))
            low = price * (1 - abs(np.random.normal(0, 0.007)))
            
            candles.append({
                'date': current_date,
                'open': price * (1 + np.random.normal(0, 0.005)),
                'high': high,
                'low': low,
                'close': price,
                'volume': int(np.random.uniform(1e6, 50e6)),
            })
            current_date += timedelta(days=1)
        
        return candles


def run_hedged_backtest():
    """Run backtest comparing hedged vs unhedged strategies"""
    
    capital = 100000
    symbols = ['META', 'AMZN', 'NFLX', 'GOOGL', 'NVDA', 'TSLA']
    
    print("\n" + "=" * 80)
    print("OPTIONS HEDGING BACKTEST")
    print("Comparing Unhedged vs Protected Put Strategy")
    print("=" * 80)
    
    # Fetch data
    price_data = {}
    for symbol in symbols:
        candles = YahooDataFetcher.get_historical_data(symbol)
        price_data[symbol] = candles
    
    spy_data = YahooDataFetcher.get_historical_data('SPY')
    
    # Create strategy
    strategy = EnhancedStockStrategy()
    
    # Create hedge manager with protective put strategy
    hedge_params = HedgeParameters(
        strategy=HedgeStrategy.PROTECTIVE_PUT,
        put_strike_pct=0.95,      # 5% below entry
        put_cost_pct=0.02,         # 2% cost
        hedge_threshold=-0.03,     # Start hedge at -3% loss
    )
    hedge_manager = OptionsHedgeManager(hedge_params)
    
    # Simulate unhedged positions
    unhedged_trades = []
    unhedged_equity = capital
    unhedged_positions = {}
    
    # Simulate hedged positions
    hedged_trades = []
    hedged_equity = capital
    hedged_positions = {}
    
    days = len(next(iter(price_data.values())))
    
    print(f"\nSimulating {days} trading days on {len(symbols)} stocks...")
    print(f"Starting Capital: ${capital:,.2f}\n")
    
    # Main backtest loop
    for day in range(50, days):
        for symbol in symbols:
            closes = [c['close'] for c in price_data[symbol][:day+1]]
            current_price = closes[-1]
            
            if len(closes) < 50:
                continue
            
            # Generate signal
            signal = strategy.analyze_stock(symbol, closes, current_price)
            
            if signal and signal.action == 'BUY':
                # UNHEDGED: Open position without protection
                if symbol not in unhedged_positions:
                    risk = current_price * 0.02
                    position_size = unhedged_equity * 0.02 / risk if risk > 0 else 0
                    
                    if position_size > 0:
                        unhedged_positions[symbol] = {
                            'entry_price': current_price,
                            'entry_day': day,
                            'size': position_size,
                            'stop_loss': current_price * 0.98,
                            'take_profit': current_price * 1.04,
                        }
                
                # HEDGED: Open position with protective put
                if symbol not in hedged_positions:
                    risk = current_price * 0.02
                    position_size = hedged_equity * 0.02 / risk if risk > 0 else 0
                    
                    if position_size > 0:
                        hedged_positions[symbol] = {
                            'entry_price': current_price,
                            'entry_day': day,
                            'size': position_size,
                            'stop_loss': current_price * 0.98,
                            'take_profit': current_price * 1.04,
                        }
                        
                        # Apply protective put hedge
                        hedge_manager.hedge_position(
                            symbol, current_price, current_price, int(position_size)
                        )
            
            # Check exits - UNHEDGED
            if symbol in unhedged_positions:
                pos = unhedged_positions[symbol]
                exit_price = None
                exit_reason = None
                
                if current_price <= pos['stop_loss']:
                    exit_price = pos['stop_loss']
                    exit_reason = "STOP_LOSS"
                elif current_price >= pos['take_profit']:
                    exit_price = pos['take_profit']
                    exit_reason = "TAKE_PROFIT"
                elif day - pos['entry_day'] >= 15:
                    exit_price = current_price
                    exit_reason = "TIME_STOP"
                
                if exit_reason:
                    pnl = (exit_price - pos['entry_price']) * pos['size']
                    unhedged_equity += pnl
                    unhedged_trades.append({'symbol': symbol, 'pnl': pnl, 'reason': exit_reason})
                    del unhedged_positions[symbol]
            
            # Check exits - HEDGED
            if symbol in hedged_positions:
                pos = hedged_positions[symbol]
                exit_price = None
                exit_reason = None
                
                unrealized_pnl_pct = (current_price - pos['entry_price']) / pos['entry_price']
                
                # Update hedge P&L
                equity_pnl = (current_price - pos['entry_price']) * pos['size']
                hedge_manager.update_hedge_pnl(symbol, current_price, equity_pnl)
                
                if current_price <= pos['stop_loss']:
                    exit_price = pos['stop_loss']
                    exit_reason = "STOP_LOSS"
                elif current_price >= pos['take_profit']:
                    exit_price = pos['take_profit']
                    exit_reason = "TAKE_PROFIT"
                elif day - pos['entry_day'] >= 15:
                    exit_price = current_price
                    exit_reason = "TIME_STOP"
                
                if exit_reason:
                    equity_pnl = (exit_price - pos['entry_price']) * pos['size']
                    
                    # Get hedge P&L
                    if symbol in hedge_manager.hedged_positions:
                        hedge = hedge_manager.hedged_positions[symbol]
                        hedge_manager.update_hedge_pnl(symbol, exit_price, equity_pnl)
                        total_pnl = equity_pnl + hedge.hedge_pnl
                        del hedge_manager.hedged_positions[symbol]
                    else:
                        total_pnl = equity_pnl
                    
                    hedged_equity += total_pnl
                    hedged_trades.append({'symbol': symbol, 'pnl': total_pnl, 'reason': exit_reason})
                    del hedged_positions[symbol]
    
    # Close remaining positions
    for symbol in list(unhedged_positions.keys()):
        pos = unhedged_positions[symbol]
        final_price = price_data[symbol][-1]['close']
        pnl = (final_price - pos['entry_price']) * pos['size']
        unhedged_equity += pnl
        unhedged_trades.append({'symbol': symbol, 'pnl': pnl, 'reason': 'EOD_CLOSE'})
    
    for symbol in list(hedged_positions.keys()):
        pos = hedged_positions[symbol]
        final_price = price_data[symbol][-1]['close']
        equity_pnl = (final_price - pos['entry_price']) * pos['size']
        
        if symbol in hedge_manager.hedged_positions:
            hedge = hedge_manager.hedged_positions[symbol]
            hedge_manager.update_hedge_pnl(symbol, final_price, equity_pnl)
            total_pnl = equity_pnl + hedge.hedge_pnl
            del hedge_manager.hedged_positions[symbol]
        else:
            total_pnl = equity_pnl
        
        hedged_equity += total_pnl
        hedged_trades.append({'symbol': symbol, 'pnl': total_pnl, 'reason': 'EOD_CLOSE'})
    
    # Calculate statistics
    def calc_stats(trades):
        if not trades:
            return {
                'count': 0, 'winners': 0, 'losers': 0, 'total_pnl': 0,
                'avg_win': 0, 'avg_loss': 0, 'profit_factor': 0
            }
        
        winners = [t['pnl'] for t in trades if t['pnl'] > 0]
        losers = [t['pnl'] for t in trades if t['pnl'] < 0]
        
        winning_sum = sum(winners) if winners else 0
        losing_sum = sum(losers) if losers else 0
        
        return {
            'count': len(trades),
            'winners': len(winners),
            'losers': len(losers),
            'total_pnl': sum([t['pnl'] for t in trades]),
            'avg_win': np.mean(winners) if winners else 0,
            'avg_loss': np.mean(losers) if losers else 0,
            'profit_factor': abs(winning_sum / losing_sum) if losing_sum != 0 else 0,
            'win_rate': len(winners) / len(trades) * 100 if trades else 0,
        }
    
    unhedged_stats = calc_stats(unhedged_trades)
    hedged_stats = calc_stats(hedged_trades)
    hedge_perf = hedge_manager.evaluate_hedge_performance()
    
    # Display results
    print("\n" + "=" * 80)
    print("BACKTEST RESULTS COMPARISON")
    print("=" * 80)
    
    print("\n📊 UNHEDGED STRATEGY")
    print(f"  Final Equity:        ${unhedged_equity:>12,.2f}")
    print(f"  Total P&L:           ${unhedged_stats['total_pnl']:>12,.2f}")
    print(f"  Return:              {(unhedged_equity-capital)/capital*100:>11.2f}%")
    print(f"  Total Trades:        {unhedged_stats['count']:>12}")
    print(f"  Win Rate:            {unhedged_stats['win_rate']:>11.1f}%")
    print(f"  Avg Win:             ${unhedged_stats['avg_win']:>12,.2f}")
    print(f"  Avg Loss:            ${unhedged_stats['avg_loss']:>12,.2f}")
    print(f"  Profit Factor:       {unhedged_stats['profit_factor']:>12.2f}x")
    
    print("\n🛡️  HEDGED STRATEGY (Protective Puts)")
    print(f"  Final Equity:        ${hedged_equity:>12,.2f}")
    print(f"  Total P&L:           ${hedged_stats['total_pnl']:>12,.2f}")
    print(f"  Return:              {(hedged_equity-capital)/capital*100:>11.2f}%")
    print(f"  Total Trades:        {hedged_stats['count']:>12}")
    print(f"  Win Rate:            {hedged_stats['win_rate']:>11.1f}%")
    print(f"  Avg Win:             ${hedged_stats['avg_win']:>12,.2f}")
    print(f"  Avg Loss:            ${hedged_stats['avg_loss']:>12,.2f}")
    print(f"  Profit Factor:       {hedged_stats['profit_factor']:>12.2f}x")
    
    print("\n💰 HEDGE PERFORMANCE")
    print(f"  Total Hedges:        {hedge_perf['total_hedges']:>12}")
    print(f"  Hedge Cost:          ${hedge_perf['total_hedge_cost']:>12,.2f}")
    print(f"  Hedge Benefit:       ${hedge_perf['total_hedge_benefit']:>12,.2f}")
    print(f"  Protected Positions: {hedge_perf['protected_positions']:>12}")
    print(f"  Loss Saved:          ${hedge_perf['loss_saved']:>12,.2f}")
    print(f"  Cost-Benefit Ratio:  {hedge_perf['cost_benefit_ratio']:>12.2f}x")
    print(f"  Hedge ROI:           {hedge_perf['roi']:>11.1f}%")
    
    print("\n📈 IMPROVEMENT (Hedged vs Unhedged)")
    pnl_diff = hedged_stats['total_pnl'] - unhedged_stats['total_pnl']
    max_loss_reduction = abs(min([t['pnl'] for t in unhedged_trades], default=0)) - \
                         abs(min([t['pnl'] for t in hedged_trades], default=0))
    
    print(f"  P&L Improvement:     ${pnl_diff:>12,.2f}")
    print(f"  Return Improvement:  {pnl_diff/capital*100:>11.2f}%")
    print(f"  Max Loss Reduction:  ${max_loss_reduction:>12,.2f}")
    
    if abs(max_loss_reduction) > abs(hedge_perf['total_hedge_cost']):
        print(f"  Status:              ✅ Hedges MORE THAN paid for themselves!")
    else:
        print(f"  Status:              ⚠️  Hedge cost exceeds max loss reduction")
    
    print("\n" + "=" * 80)
    
    # Export to Excel
    if EXCEL_AVAILABLE:
        try:
            export_comparison_to_excel(
                unhedged_trades, hedged_trades, 
                unhedged_stats, hedged_stats, hedge_perf,
                capital
            )
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
    
    return {
        'unhedged': unhedged_stats,
        'hedged': hedged_stats,
        'hedge_perf': hedge_perf,
        'improvement': {
            'pnl_diff': pnl_diff,
            'return_diff': pnl_diff / capital * 100,
            'max_loss_reduction': max_loss_reduction,
        }
    }


def export_comparison_to_excel(unhedged_trades, hedged_trades, 
                               unhedged_stats, hedged_stats, hedge_perf,
                               capital):
    """Export hedging comparison to Excel"""
    
    filename = 'hedging_backtest_comparison.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comparison'
    
    # Summary comparison
    row = 1
    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Unhedged'
    ws[f'C{row}'] = 'Hedged'
    ws[f'D{row}'] = 'Difference'
    row += 1
    
    metrics = [
        ('Total Trades', unhedged_stats['count'], hedged_stats['count']),
        ('Total P&L', unhedged_stats['total_pnl'], hedged_stats['total_pnl']),
        ('Return %', (unhedged_stats['total_pnl']/capital)*100, (hedged_stats['total_pnl']/capital)*100),
        ('Win Rate %', unhedged_stats['win_rate'], hedged_stats['win_rate']),
        ('Average Win', unhedged_stats['avg_win'], hedged_stats['avg_win']),
        ('Average Loss', unhedged_stats['avg_loss'], hedged_stats['avg_loss']),
        ('Profit Factor', unhedged_stats['profit_factor'], hedged_stats['profit_factor']),
    ]
    
    for metric, unhedged_val, hedged_val in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = f"{unhedged_val:.2f}" if isinstance(unhedged_val, float) else unhedged_val
        ws[f'C{row}'] = f"{hedged_val:.2f}" if isinstance(hedged_val, float) else hedged_val
        ws[f'D{row}'] = f"{hedged_val - unhedged_val:.2f}" if isinstance(hedged_val, float) else hedged_val - unhedged_val
        row += 1
    
    wb.save(filename)
    logger.info(f"Excel report saved: {filename}")
    print(f"\n✓ Excel report saved: {filename}")


if __name__ == '__main__':
    try:
        results = run_hedged_backtest()
        sys.exit(0)
    except Exception as e:
        logger.error(f"Backtest failed: {e}", exc_info=True)
        print(f"\n❌ Backtest failed: {e}")
        sys.exit(1)
