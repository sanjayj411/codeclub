#!/usr/bin/env python3
"""
FANG + NVDA + TSLA 1-year backtest with Excel report generation
Stocks: META (Facebook), AMZN (Amazon), NFLX (Netflix), GOOGL (Google), NVDA, TSLA (Tesla)
"""

import sys
import json
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
import numpy as np

from src.backtesting.backtester import Backtester, Trade
from src.strategy.trading_strategy import TradingStrategy
from src.core.logger import logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available - install with: pip install openpyxl")


class YahooDataFetcher:
    """Fetch historical data - defaults to realistic sample data"""
    
    @staticmethod
    def get_historical_data(symbol: str, days: int = 365) -> List[Dict]:
        """
        Get historical price data for a symbol
        
        Args:
            symbol: Stock ticker (e.g., 'AAPL')
            days: Number of days of history to fetch
            
        Returns:
            List of candles with OHLCV data
        """
        logger.info(f"Generating realistic sample data for {symbol} ({days} days)")
        return YahooDataFetcher._generate_sample_data(symbol, days)
    
    @staticmethod
    def _generate_sample_data(symbol: str, days: int = 365) -> List[Dict]:
        """Generate realistic sample data (fallback)"""
        np.random.seed(hash(symbol) % 2**32)
        
        # Base prices for stocks
        base_prices = {
            'META': 150.0,
            'AMZN': 170.0,
            'NFLX': 250.0,
            'GOOGL': 140.0,
            'NVDA': 850.0,
            'TSLA': 280.0
        }
        
        price = base_prices.get(symbol, 100.0)
        candles = []
        
        end_date = datetime.now()
        current_date = end_date - timedelta(days=days)
        
        for i in range(days):
            # Random walk with drift
            daily_return = np.random.normal(0.0005, 0.02)  # 0.05% mean, 2% volatility
            price = price * (1 + daily_return)
            
            open_p = price * (1 + np.random.normal(0, 0.005))
            high = max(open_p, price) * (1 + abs(np.random.normal(0, 0.01)))
            low = min(open_p, price) * (1 - abs(np.random.normal(0, 0.01)))
            close = price
            volume = np.random.uniform(20000000, 100000000)
            
            candle = {
                'timestamp': current_date,
                'open': float(open_p),
                'high': float(high),
                'low': float(low),
                'close': float(close),
                'volume': float(volume)
            }
            candles.append(candle)
            current_date += timedelta(days=1)
        
        logger.info(f"Generated {len(candles)} sample candles for {symbol}")
        return candles


class FANGBacktester:
    """Run backtest for FANG + NVDA + TSLA stocks with SPY market filter"""
    
    def __init__(self, initial_capital: float = 100000.0):
        """Initialize FANG backtester"""
        self.symbols = ['META', 'AMZN', 'NFLX', 'GOOGL', 'NVDA', 'TSLA']
        self.initial_capital = initial_capital
        # Use standard TradingStrategy with RSI + MACD
        self.strategy = TradingStrategy(
            rsi_period=14,
            rsi_oversold=30,
            rsi_overbought=70,
            min_drawdown_for_buy=2.0
        )
        self.backtester = Backtester(self.strategy, initial_capital)
        self.spy_market_filter = 2.0  # Only trade when SPY is 2%+ down
        
        self.results: Dict = {
            'metadata': {},
            'stats_by_symbol': {},
            'all_trades': [],
            'summary_stats': {}
        }
    
    def run(self, days: int = 365) -> Dict:
        """
        Run backtest for 1 year with SPY market filter
        
        Args:
            days: Number of days to backtest (default 365)
            
        Returns:
            Results dictionary
        """
        logger.info(f"Starting {', '.join(self.symbols)} backtest for {days} days")
        logger.info(f"Market filter: Only trade when SPY is {self.spy_market_filter}%+ down")
        logger.info(f"Initial capital: ${self.initial_capital:,.2f}")
        
        # Fetch data for all symbols + SPY for market filter
        price_data = {}
        symbols_to_fetch = self.symbols + ['SPY']
        
        for symbol in symbols_to_fetch:
            logger.info(f"Fetching data for {symbol}...")
            candles = YahooDataFetcher.get_historical_data(symbol, days)
            if candles:
                price_data[symbol] = candles
        
        if not price_data:
            logger.error("No price data fetched")
            return self.results
        
        # Run backtest with market filter
        logger.info("Running backtest with SPY market filter...")
        stats = self._run_with_spy_filter(price_data, days)
        
        # Collect results
        self.results['metadata'] = {
            'backtest_date': datetime.now().isoformat(),
            'symbols': self.symbols,
            'days': days,
            'initial_capital': self.initial_capital,
            'strategy': 'RSI + MACD with SPY 2% down market filter',
            'market_filter': f'SPY {self.spy_market_filter}% drawdown',
        }
        
        self.results['summary_stats'] = {
            'total_trades': stats.total_trades,
            'winning_trades': stats.winning_trades,
            'losing_trades': stats.losing_trades,
            'win_rate': stats.win_rate,
            'total_pnl': stats.total_pnl,
            'total_pnl_pct': stats.total_pnl_pct,
            'avg_win': stats.avg_win,
            'avg_loss': stats.avg_loss,
            'profit_factor': stats.profit_factor,
            'max_drawdown': stats.max_drawdown,
            'max_drawdown_pct': stats.max_drawdown_pct,
            'sharpe_ratio': stats.sharpe_ratio,
        }
        
        # Group trades by symbol
        trades_by_symbol = {}
        for trade in stats.trades:
            if trade.symbol not in trades_by_symbol:
                trades_by_symbol[trade.symbol] = []
            trades_by_symbol[trade.symbol].append(trade)
        
        # Calculate stats per symbol
        for symbol in self.symbols:
            trades = trades_by_symbol.get(symbol, [])
            self.results['stats_by_symbol'][symbol] = self._calculate_symbol_stats(symbol, trades)
            self.results['all_trades'].extend([self._trade_to_dict(t) for t in trades])
        
        logger.info(f"Backtest completed: {stats.total_trades} total trades")
        logger.info(f"Total P&L: ${stats.total_pnl:,.2f} ({stats.total_pnl_pct:.2f}%)")
        logger.info(f"Win rate: {stats.win_rate:.1f}%")
        
        return self.results
    
    def _run_with_spy_filter(self, price_data: Dict, days: int) -> any:
        """Run backtest with SPY market filter - only trade when SPY is 2%+ down"""
        from src.backtesting.backtester import BacktestStats
        
        logger.info(f"Processing backtest with SPY market filter ({self.spy_market_filter}% down)")
        
        # Reset backtester state
        self.backtester.positions = {}
        self.backtester.trades = []
        self.backtester.orders = []
        self.backtester.equity_curve = []
        self.backtester.capital = self.initial_capital
        
        # Get all timestamps across all symbols and sort them
        all_timestamps = set()
        for symbol, candles in price_data.items():
            for candle in candles:
                all_timestamps.add(candle['timestamp'])
        
        timestamps = sorted(all_timestamps)
        
        # Process each timestamp
        for timestamp in timestamps:
            # Get current prices for all positions
            current_prices = {}
            closes_by_symbol = {symbol: [] for symbol in self.symbols}  # Only track main symbols
            spy_closes = []
            
            # Calculate SPY drawdown for market filter
            spy_candles = price_data.get('SPY', [])
            spy_candles_up_to = [c for c in spy_candles if c['timestamp'] <= timestamp]
            if spy_candles_up_to:
                spy_current = spy_candles_up_to[-1]['close']
                spy_closes = [c['close'] for c in spy_candles_up_to]
                if len(spy_closes) >= 20:
                    spy_high_20 = max(spy_closes[-20:])
                    spy_drawdown = ((spy_high_20 - spy_current) / spy_high_20) * 100
                else:
                    spy_drawdown = 0
            else:
                spy_drawdown = 0
            
            # Check if SPY meets market filter (2%+ down)
            market_allows_trading = spy_drawdown >= self.spy_market_filter
            
            for symbol, candles in price_data.items():
                if symbol == 'SPY':
                    continue  # Skip SPY, it's only for filtering
                    
                candles_up_to = [c for c in candles if c['timestamp'] <= timestamp]
                if candles_up_to:
                    current_prices[symbol] = candles_up_to[-1]['close']
                    closes_by_symbol[symbol] = [c['close'] for c in candles_up_to]
            
            # Update position values
            portfolio_value = self.backtester.capital
            for symbol, position in self.backtester.positions.items():
                if symbol in current_prices:
                    position.current_price = current_prices[symbol]
                    portfolio_value += position.value
            
            # Record equity
            self.backtester.equity_curve.append((timestamp, portfolio_value))
            
            # Analyze each symbol for trading signals ONLY if SPY meets market filter
            if market_allows_trading:
                for symbol, closes in closes_by_symbol.items():
                    if not closes or len(closes) < 30:
                        continue
                    
                    if symbol in current_prices:
                        current_price = current_prices[symbol]
                        
                        # Get trading signal
                        signal = self.backtester.strategy.analyze(symbol, closes, current_price)
                        
                        # Execute trades based on signal
                        self.backtester._execute_signal(signal, timestamp)
        
        # Close all remaining positions
        final_timestamp = timestamps[-1] if timestamps else datetime.now()
        for symbol in list(self.backtester.positions.keys()):
            if symbol in current_prices:
                self.backtester._close_position(symbol, current_prices[symbol], final_timestamp)
        
        # Calculate statistics
        stats = self.backtester._calculate_stats()
        return stats
    
    def _calculate_symbol_stats(self, symbol: str, trades: List[Trade]) -> Dict:
        """Calculate statistics for a single symbol"""
        if not trades:
            return {
                'symbol': symbol,
                'total_trades': 0,
                'winning_trades': 0,
                'losing_trades': 0,
                'win_rate': 0.0,
                'total_pnl': 0.0,
                'total_pnl_pct': 0.0,
                'avg_win': 0.0,
                'avg_loss': 0.0,
                'profit_factor': 0.0,
            }
        
        total_pnl = sum(t.pnl for t in trades)
        winning = [t for t in trades if t.pnl > 0]
        losing = [t for t in trades if t.pnl < 0]
        
        win_rate = (len(winning) / len(trades) * 100) if trades else 0
        avg_win = sum(t.pnl for t in winning) / len(winning) if winning else 0
        avg_loss = sum(t.pnl for t in losing) / len(losing) if losing else 0
        
        total_win = sum(t.pnl for t in winning)
        total_loss = abs(sum(t.pnl for t in losing))
        profit_factor = total_win / total_loss if total_loss > 0 else (1 if total_win > 0 else 0)
        
        first_entry = min(t.entry_price * t.quantity for t in trades)
        total_pnl_pct = (total_pnl / first_entry * 100) if first_entry > 0 else 0
        
        return {
            'symbol': symbol,
            'total_trades': len(trades),
            'winning_trades': len(winning),
            'losing_trades': len(losing),
            'win_rate': win_rate,
            'total_pnl': total_pnl,
            'total_pnl_pct': total_pnl_pct,
            'avg_win': avg_win,
            'avg_loss': avg_loss,
            'profit_factor': profit_factor,
        }
    
    def _trade_to_dict(self, trade: Trade) -> Dict:
        """Convert Trade object to dictionary"""
        return {
            'symbol': trade.symbol,
            'entry_price': trade.entry_price,
            'exit_price': trade.exit_price,
            'quantity': trade.quantity,
            'entry_time': trade.entry_time.isoformat(),
            'exit_time': trade.exit_time.isoformat(),
            'pnl': trade.pnl,
            'pnl_pct': trade.pnl_pct,
            'days_held': (trade.exit_time - trade.entry_time).days,
        }
    
    def export_to_excel(self, filename: str = 'fang_nvda_backtest.xlsx') -> str:
        """
        Export results to Excel with detailed analysis
        
        Args:
            filename: Output Excel filename
            
        Returns:
            Path to generated file
        """
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl required for Excel export")
            return ""
        
        logger.info(f"Creating Excel report: {filename}")
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Add sheets
        self._add_summary_sheet(wb)
        self._add_symbol_stats_sheet(wb)
        self._add_trades_sheet(wb)
        self._add_detailed_analysis_sheet(wb)
        
        # Save
        wb.save(filename)
        logger.info(f"Excel report saved: {filename}")
        
        return filename
    
    def _add_summary_sheet(self, wb):
        """Add summary sheet to workbook"""
        ws = wb.create_sheet('Summary', 0)
        
        # Title
        ws['A1'] = 'FANG + NVDA 1-Year Backtest Report'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Metadata
        row = 3
        ws[f'A{row}'] = 'Backtest Date:'
        ws[f'B{row}'] = self.results['metadata']['backtest_date']
        row += 1
        
        ws[f'A{row}'] = 'Symbols:'
        ws[f'B{row}'] = ', '.join(self.results['metadata']['symbols'])
        row += 1
        
        ws[f'A{row}'] = 'Period:'
        ws[f'B{row}'] = f"{self.results['metadata']['days']} days"
        row += 1
        
        ws[f'A{row}'] = 'Initial Capital:'
        ws[f'B{row}'] = self.results['metadata']['initial_capital']
        ws[f'B{row}'].number_format = '$#,##0.00'
        row += 2
        
        # Summary Statistics
        ws[f'A{row}'] = 'SUMMARY STATISTICS'
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        
        stats = self.results['summary_stats']
        
        summary_data = [
            ('Total Trades', stats['total_trades'], '0'),
            ('Winning Trades', stats['winning_trades'], '0'),
            ('Losing Trades', stats['losing_trades'], '0'),
            ('Win Rate', stats['win_rate'], '0.00"%"'),
            ('Total P&L', stats['total_pnl'], '$#,##0.00'),
            ('Total Return %', stats['total_pnl_pct'], '0.00"%"'),
            ('Average Win', stats['avg_win'], '$#,##0.00'),
            ('Average Loss', stats['avg_loss'], '$#,##0.00'),
            ('Profit Factor', stats['profit_factor'], '0.00'),
            ('Max Drawdown', stats['max_drawdown'], '$#,##0.00'),
            ('Max Drawdown %', stats['max_drawdown_pct'], '0.00"%"'),
            ('Sharpe Ratio', stats['sharpe_ratio'], '0.00'),
        ]
        
        for label, value, fmt in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if fmt:
                ws[f'B{row}'].number_format = fmt
            row += 1
        
        # Style
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=2):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    
    def _add_symbol_stats_sheet(self, wb):
        """Add symbol statistics sheet"""
        ws = wb.create_sheet('Symbol Stats', 1)
        
        # Headers
        headers = ['Symbol', 'Trades', 'Wins', 'Losses', 'Win Rate', 'Total P&L', 'Return %', 'Avg Win', 'Avg Loss', 'Profit Factor']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = 2
        for symbol in self.symbols:
            stats = self.results['stats_by_symbol'].get(symbol, {})
            
            data = [
                stats.get('symbol', symbol),
                stats.get('total_trades', 0),
                stats.get('winning_trades', 0),
                stats.get('losing_trades', 0),
                stats.get('win_rate', 0),
                stats.get('total_pnl', 0),
                stats.get('total_pnl_pct', 0),
                stats.get('avg_win', 0),
                stats.get('avg_loss', 0),
                stats.get('profit_factor', 0),
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                # Format numbers
                if col >= 2:  # Number columns
                    if col == 5:  # Win Rate
                        cell.number_format = '0.00"%"'
                    elif col in [6, 8, 9]:  # P&L columns
                        cell.number_format = '$#,##0.00'
                    elif col == 7:  # Return %
                        cell.number_format = '0.00"%"'
                    else:
                        cell.number_format = '0'
                
                # Color code P&L
                if col == 6:  # Total P&L
                    if value > 0:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif value < 0:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _add_trades_sheet(self, wb):
        """Add detailed trades sheet"""
        ws = wb.create_sheet('All Trades', 2)
        
        # Headers
        headers = ['Symbol', 'Entry Date', 'Entry Price', 'Exit Date', 'Exit Price', 'Quantity', 'P&L', 'P&L %', 'Days Held']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = 2
        for trade in sorted(self.results['all_trades'], key=lambda t: t['entry_time']):
            data = [
                trade['symbol'],
                datetime.fromisoformat(trade['entry_time']).strftime('%Y-%m-%d'),
                trade['entry_price'],
                datetime.fromisoformat(trade['exit_time']).strftime('%Y-%m-%d'),
                trade['exit_price'],
                trade['quantity'],
                trade['pnl'],
                trade['pnl_pct'],
                trade['days_held'],
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                # Format
                if col in [3, 5]:  # Prices
                    cell.number_format = '$#,##0.00'
                elif col == 6:  # Quantity
                    cell.number_format = '0.0000'
                elif col == 7:  # P&L
                    cell.number_format = '$#,##0.00'
                    # Color code
                    if value > 0:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif value < 0:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif col == 8:  # P&L %
                    cell.number_format = '0.00"%"'
                else:
                    cell.alignment = Alignment(horizontal='center')
                
                if col in [2, 4]:  # Dates
                    cell.alignment = Alignment(horizontal='center')
            
            row += 1
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _add_detailed_analysis_sheet(self, wb):
        """Add detailed analysis sheet"""
        ws = wb.create_sheet('Analysis', 3)
        
        ws['A1'] = 'DETAILED ANALYSIS'
        ws['A1'].font = Font(size=12, bold=True)
        
        row = 3
        
        # Performance by symbol
        ws[f'A{row}'] = 'Performance by Symbol'
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        
        for symbol in self.symbols:
            stats = self.results['stats_by_symbol'].get(symbol, {})
            if stats.get('total_trades', 0) > 0:
                ws[f'A{row}'] = f"{symbol}:"
                ws[f'B{row}'] = f"{stats['total_trades']} trades, {stats['win_rate']:.1f}% wins, {stats['total_pnl']:,.2f} P&L"
                row += 1
        
        row += 2
        
        # Best and worst trades
        ws[f'A{row}'] = 'Best & Worst Trades'
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        
        trades = self.results['all_trades']
        if trades:
            best_trade = max(trades, key=lambda t: t['pnl'])
            worst_trade = min(trades, key=lambda t: t['pnl'])
            
            ws[f'A{row}'] = 'Best Trade:'
            ws[f'B{row}'] = f"{best_trade['symbol']} +${best_trade['pnl']:.2f} ({best_trade['pnl_pct']:.2f}%)"
            row += 1
            
            ws[f'A{row}'] = 'Worst Trade:'
            ws[f'B{row}'] = f"{worst_trade['symbol']} ${worst_trade['pnl']:.2f} ({worst_trade['pnl_pct']:.2f}%)"
            row += 1
        
        row += 2
        
        # Insights
        ws[f'A{row}'] = 'Key Insights'
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        
        stats = self.results['summary_stats']
        
        insights = [
            f"Total Return: {stats['total_pnl_pct']:.2f}% on initial ${self.initial_capital:,.2f}",
            f"Strategy generated {stats['total_trades']} trades over {self.results['metadata']['days']} days",
            f"Win rate: {stats['win_rate']:.1f}% with avg win/loss ratio: {stats['avg_win'] / abs(stats['avg_loss']) if stats['avg_loss'] != 0 else 0:.2f}",
            f"Risk (Max Drawdown): ${stats['max_drawdown']:,.2f} ({stats['max_drawdown_pct']:.2f}%)",
            f"Sharpe Ratio: {stats['sharpe_ratio']:.2f} (higher is better)",
            f"Best performing stock: {max(self.symbols, key=lambda s: self.results['stats_by_symbol'].get(s, {}).get('total_pnl', 0))}",
        ]
        
        for insight in insights:
            ws[f'A{row}'] = insight
            row += 1
        
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 40


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='FANG + NVDA + TSLA 1-year backtest')
    parser.add_argument('--capital', type=float, default=100000, help='Initial capital (default: $100,000)')
    parser.add_argument('--days', type=int, default=365, help='Backtest days (default: 365)')
    parser.add_argument('--output', type=str, default='fang_nvda_backtest.xlsx', help='Output Excel file')
    parser.add_argument('--json', type=str, help='Also export to JSON')
    
    args = parser.parse_args()
    
    # Run backtest
    backtester = FANGBacktester(initial_capital=args.capital)
    results = backtester.run(days=args.days)
    
    # Print summary
    print("\n" + "="*70)
    print("FANG + NVDA + TSLA BACKTEST RESULTS".center(70))
    print("="*70)
    
    stats = results['summary_stats']
    print(f"\nInitial Capital:        ${args.capital:>15,.2f}")
    print(f"Total Trades:           {stats['total_trades']:>15}")
    print(f"Winning Trades:         {stats['winning_trades']:>15} ({stats['win_rate']:.1f}%)")
    print(f"Losing Trades:          {stats['losing_trades']:>15}")
    print(f"\nTotal P&L:              ${stats['total_pnl']:>15,.2f}")
    print(f"Total Return:           {stats['total_pnl_pct']:>15,.2f}%")
    print(f"Average Win:            ${stats['avg_win']:>15,.2f}")
    print(f"Average Loss:           ${stats['avg_loss']:>15,.2f}")
    print(f"Profit Factor:          {stats['profit_factor']:>15.2f}x")
    print(f"\nMax Drawdown:           ${stats['max_drawdown']:>15,.2f} ({stats['max_drawdown_pct']:.2f}%)")
    print(f"Sharpe Ratio:           {stats['sharpe_ratio']:>15.2f}")
    
    print("\n" + "-"*70)
    print("Symbol Performance:".ljust(70))
    print("-"*70)
    
    for symbol in backtester.symbols:
        sym_stats = results['stats_by_symbol'].get(symbol, {})
        if sym_stats.get('total_trades', 0) > 0:
            print(f"{symbol:6} | {sym_stats['total_trades']:3} trades | "
                  f"{sym_stats['win_rate']:5.1f}% wins | "
                  f"P&L: ${sym_stats['total_pnl']:>10,.2f} ({sym_stats['total_pnl_pct']:6.2f}%)")
    
    # Export to Excel
    if EXCEL_AVAILABLE:
        excel_file = backtester.export_to_excel(args.output)
        print(f"\n✓ Excel report saved: {excel_file}")
    else:
        print("\n⚠ Excel export requires openpyxl: pip install openpyxl")
    
    # Export to JSON if requested
    if args.json:
        with open(args.json, 'w') as f:
            json.dump(results, f, indent=2, default=str)
        print(f"✓ JSON report saved: {args.json}")
    
    print("\n" + "="*70)


if __name__ == '__main__':
    main()
